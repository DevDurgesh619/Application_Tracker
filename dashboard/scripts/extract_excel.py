import openpyxl, json, re
wb = openpyxl.load_workbook('Mehek_College_Tracker_2027_VERIFIED.xlsx', data_only=True)

SOURCE_FIXES = {
    "https://admission.universityofcalifornia.edu/admission-requirements/freshman-requirements/standardized-testing.html": "https://admission.universityofcalifornia.edu/admission-requirements/freshman-requirements/",
    "https://admission.universityofcalifornia.edu/tuition-financial-aid/types-of-aid/who-can-get-financial-aid.html": "https://admission.universityofcalifornia.edu/tuition-financial-aid/types-of-aid/who-can-get-financial-aid",
    "https://admissions.cornell.edu/how-apply/testing": "https://admissions.cornell.edu/",
    "https://admissions.duke.edu/first-year-applicants/admissions-statistics/": "https://admissions.duke.edu/",
    "https://admissions.duke.edu/first-year-applicants/application-deadlines-and-decisions/": "https://admissions.duke.edu/",
    "https://admissions.duke.edu/first-year-applicants/standardized-testing/": "https://admissions.duke.edu/",
    "https://admissions.upenn.edu/admissions-and-financial-aid/what-penn-looks-for/testing-policy": "https://admissions.upenn.edu/admissions-and-financial-aid/what-penn-looks-for/",
    "https://admissions.upenn.edu/admissions-and-financial-aid/whats-it-take/how-apply": "https://admissions.upenn.edu/admissions-and-financial-aid/",
    "https://admissions.yale.edu/application-timeline-options": "https://admissions.yale.edu/",
    "https://admissions.yale.edu/class-profile": "https://admissions.yale.edu/",
    "https://admissions.yale.edu/financial-aid-international-students": "https://admissions.yale.edu/",
    "https://apply.emory.edu/apply/deadlines.html": "https://apply.emory.edu/apply/",
    "https://apply.emory.edu/apply/testing-policy.html": "https://apply.emory.edu/apply/",
    "https://apply.emory.edu/discover/class-profile.html": "https://apply.emory.edu/discover/",
    "https://apply.emory.edu/discover/cost-and-aid.html": "https://apply.emory.edu/discover/",
    "https://apply.emory.edu/discover/cost-and-aid/international.html": "https://apply.emory.edu/discover/",
    "https://econ.duke.edu/undergraduate/majors-and-minor/bs-degree": "https://econ.duke.edu/undergraduate/",
    "https://economics.emory.edu/home/undergraduate/stem.html": "https://economics.emory.edu/",
    "https://economics.uci.edu/undergraduate/": "https://economics.uci.edu/",
    "https://economics.ucla.edu/stem/": "https://economics.ucla.edu/",
    "https://economics.ucsd.edu/undergraduate-program/major-requirements/index.html": "https://economics.ucsd.edu/undergraduate-program/",
    "https://finaid.cornell.edu/types-aid/international-students": "https://finaid.cornell.edu/types-aid/",
    "https://finaid.yale.edu/cost-financial-aid/yale-cost": "https://finaid.yale.edu/",
    "https://financialaid.duke.edu/undergraduate-aid/cost-of-attendance/": "https://financialaid.duke.edu/",
    "https://oiss.ucsb.edu/services/employment/stem-opt": "https://oiss.ucsb.edu/",
    "https://srfs.upenn.edu/affordability/international-students": "https://srfs.upenn.edu/",
    "https://srfs.upenn.edu/cost-of-attendance": "https://srfs.upenn.edu/",
    "https://www.ntu.edu.sg/sss/admissions/undergraduate": "https://www.ntu.edu.sg/sss/admissions/",
    "https://www.unsw.edu.au/study/undergraduate/scholarships": "https://www.unsw.edu.au/study/undergraduate/",
}

def rows_of(name):
    return [list(r) for r in wb[name].iter_rows(values_only=True)]

def clean(v):
    if v is None: return None
    if isinstance(v,str):
        return v.strip()
    return v

def slug(name):
    s = name.lower()
    s = s.replace('–','-').replace('—','-')
    s = re.sub(r'[^a-z0-9]+','-',s).strip('-')
    return s

# ---------- College List ----------
cl = rows_of('📋 College List')
hdr_idx = next(i for i,r in enumerate(cl) if r and r[0]=='#')
cols = [str(c).strip() for c in cl[hdr_idx]]
unis = []
for r in cl[hdr_idx+1:]:
    if not (r[0] is not None and isinstance(r[0],(int,float))): continue
    d = {cols[i]: clean(r[i]) for i in range(len(cols))}
    unis.append(d)

# ---------- Cost Breakdown (numeric cols only, by name) ----------
cb = rows_of('💰 Cost Breakdown')
chdr = next(i for i,r in enumerate(cb) if r and r[0]=='#')
ccols = [str(c).strip() for c in cb[chdr]]
cost_by_name = {}
for r in cb[chdr+1:]:
    if not (r[0] is not None and isinstance(r[0],(int,float))): continue
    d = {ccols[i]: clean(r[i]) for i in range(len(ccols))}
    cost_by_name[d['University']] = d

# ---------- SAT & Stats ----------
ss = rows_of('📊 SAT & Stats')
shdr = next(i for i,r in enumerate(ss) if r and r[0]=='#')
scols = [str(c).strip() for c in ss[shdr]]
sat_by_name = {}
for r in ss[shdr+1:]:
    if not (r[0] is not None and isinstance(r[0],(int,float))): continue
    d = {scols[i]: clean(r[i]) for i in range(len(scols))}
    sat_by_name[d['University']] = d

# ---------- Links (with embedded hyperlinks) ----------
wbf = openpyxl.load_workbook('Mehek_College_Tracker_2027_VERIFIED.xlsx')
lws = wbf['🔗 Links']
lrows = list(lws.iter_rows())
# header row index where col B == 'University'
links_by_name = {}
for row in lrows:
    name_cell = row[1] if len(row)>1 else None
    if not name_cell or not isinstance(name_cell.value,str): continue
    nm = name_cell.value.strip()
    if nm in ('University','LINKS REFERENCE  ·  All Apply / Course / Aid Links'): continue
    apply_l = row[2].hyperlink.target if len(row)>2 and row[2].hyperlink else None
    course_l = row[3].hyperlink.target if len(row)>3 and row[3].hyperlink else None
    aid_l = row[4].hyperlink.target if len(row)>4 and row[4].hyperlink else None
    note = row[5].value if len(row)>5 else None
    links_by_name[nm] = {'apply':apply_l,'course':course_l,'aid':aid_l,'note':clean(note)}

# ---------- Essay Tracker ----------
et = rows_of('✍️ Essay Tracker')
ehdr = next(i for i,r in enumerate(et) if r and r[0]=='University')
ecols = [str(c).strip() for c in et[ehdr]]
essays = []
for r in et[ehdr+1:]:
    if not r[0]: continue
    d = {ecols[i]: clean(r[i]) for i in range(len(ecols))}
    essays.append(d)

# ---------- Interview Log ----------
il = rows_of('🎙️ Interview Log')
ihdr = next(i for i,r in enumerate(il) if r and r[0]=='University')
icols = [str(c).strip() for c in il[ihdr]]
interviews = []
for r in il[ihdr+1:]:
    if not r[0]: continue
    d = {icols[i]: clean(r[i]) for i in range(len(icols))}
    interviews.append(d)

# ---------- Activities & Honors ----------
al = rows_of('📌 Activity List')
# Activities header row: 'Activity / Honor Name'
ahdr = next(i for i,r in enumerate(al) if r and r[0] and 'Activity / Honor Name' in str(r[0]))
acols = [str(c).strip() for c in al[ahdr]]
honors_hdr = next(i for i,r in enumerate(al) if r and r[0] and 'Honor / Award Name' in str(r[0]))
activities=[]
for r in al[ahdr+1:honors_hdr]:
    if not r[0] or 'HONORS' in str(r[0]): continue
    activities.append({acols[i]: clean(r[i]) for i in range(len(acols))})
hcols = [str(c).strip() for c in al[honors_hdr]]
honors=[]
for r in al[honors_hdr+1:]:
    if not r[0]: continue
    honors.append({hcols[i]: clean(r[i]) for i in range(len(hcols))})

# ---------- Verification Audit ----------
va = rows_of('🔍 Verification Audit')
vhdr = next(i for i,r in enumerate(va) if r and r[0]=='#')
vcols = [str(c).strip() for c in va[vhdr]]
audit=[]
for r in va[vhdr+1:]:
    if not (r[0] is not None and isinstance(r[0],(int,float))): continue
    d = {vcols[i]: clean(r[i]) for i in range(len(vcols))}
    if d.get('Official source') in SOURCE_FIXES:
        d['Official source'] = SOURCE_FIXES[d['Official source']]
    audit.append(d)

# ---------- Read-Me "biggest errors" ----------
rm = rows_of('✅ Verified - Read Me')
big_errors=[]
capture=False
for r in rm:
    first = clean(r[0]) if r else None
    if first and 'BIGGEST ERRORS' in str(first): capture=True; continue
    if capture:
        if first and str(first).startswith('See the'): break
        if first and str(first).startswith('-'):
            big_errors.append(str(first).lstrip('- ').strip())

# ---------- assemble master ----------
master=[]
for u in unis:
    name = u['University']
    entry = dict(u)
    entry['id'] = slug(name)
    entry['cost'] = cost_by_name.get(name)
    entry['sat'] = sat_by_name.get(name)
    entry['links'] = links_by_name.get(name)
    entry['essays'] = [e for e in essays if e.get('University') and (e['University']==name or name.split(' ')[0] in str(e['University']))]
    entry['interviews'] = [iv for iv in interviews if iv.get('University')==name]
    master.append(entry)

out = {
  'student': {'name':'Mehek Jain','classOf':2027,'profile':'IB 40-43 Predicted · STEM-Economics','schools':len(unis)},
  'universities': master,
  'essays': essays,
  'interviews': interviews,
  'activities': activities,
  'honors': honors,
  'audit': audit,
  'bigErrors': big_errors,
}
print('Unis:',len(master),'Essays:',len(essays),'Interviews:',len(interviews),'Activities:',len(activities),'Honors:',len(honors),'Audit:',len(audit),'BigErrors:',len(big_errors))
print('Links matched:', sum(1 for u in master if u['links']))
print('Cost matched:', sum(1 for u in master if u['cost']))
print('SAT matched:', sum(1 for u in master if u['sat']))
import os
os.makedirs('/tmp/out',exist_ok=True)
json.dump(out, open('/tmp/out/master.json','w'), ensure_ascii=False, indent=2)
print('written /tmp/out/master.json')
